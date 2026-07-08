"""DATA_QUERY 表格解析与动态表导入计划。

DATA_QUERY 的 Excel/CSV 不走文档检索链路，而是被解释为结构化数据表。
本模块只负责三件事：
1. 把上传文件解析成“一个表”的内存结构；
2. 生成后端控制的物理表名、物理列名和元数据；
3. 将导入计划交给 repository 在同一个数据库事务中执行。
"""

from __future__ import annotations

import csv
import logging
import time
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Any

from openpyxl import load_workbook

from app.domains.document.components.data_query_identifiers import (
    build_data_query_physical_table_name,
    quote_generated_identifier,
)
from app.domains.document.shared.file_types import DocumentFileType
from app.domains.document.shared.errors import DataQueryIngestionFailed
from app.domains.document.components.storage_keys import original_object_key


logger = logging.getLogger(__name__)


class DataQuerySpreadsheetInvalid(Exception):
    """DATA_QUERY 表格无法解析为唯一数据表时抛出。"""


@dataclass(frozen=True, slots=True)
class ParsedDataQueryTable:
    """已解析出的单表数据。

    `original_sheet_name` 用于保留 Excel sheet 名或 CSV 的虚拟 sheet 名；
    `headers` 是用户上传的原始表头，只写入 metadata；
    `rows` 是已经按表头宽度补齐并转成字符串的行数据。
    """

    original_sheet_name: str
    headers: list[str]
    rows: list[list[str]]


@dataclass(frozen=True, slots=True)
class DataQueryTablePlan:
    """动态表导入计划。

    该对象是解析层和持久化层之间的边界：identifier、DDL、columns_info 和行数据
    都在这里成型，repository 只负责在事务中检查状态并执行这些计划。
    """

    physical_table_name: str
    column_names: list[str]
    create_sql: str
    columns_info: dict[str, Any]
    rows: list[list[str]]


def parse_data_query_spreadsheet(
    *,
    file_type: str,
    content: bytes,
    filename: str = "",
) -> ParsedDataQueryTable:
    """将 Excel 或 CSV 字节解析为唯一的表形数据集。

    Excel 必须只有一个包含表头和数据行的有效 sheet；完全空 sheet 会被忽略。
    CSV 天然只有一个表，虚拟 sheet 名固定为 `Data`，便于后续 metadata 结构统一。
    """

    if file_type == DocumentFileType.EXCEL.value:
        return _parse_excel(content=content, filename=filename)
    if file_type == DocumentFileType.CSV.value:
        return _parse_csv(content)
    raise DataQuerySpreadsheetInvalid("unsupported spreadsheet file type")


async def ingest_data_query_spreadsheet_document(
    *,
    document: Any,
    document_repository: Any,
    storage: Any,
) -> None:
    """把一个 DATA_QUERY spreadsheet 文档导入到生成的 PostgreSQL 物理表。

    这里不直接操作事务，而是完成 worker 侧的编排：
    1. 读取上传阶段预留的 `table_meta`；
    2. 从对象存储下载原始文件；
    3. 解析单表数据并构建动态表导入计划；
    4. 委托 repository 在数据库事务中创建表、插入数据并推进文档状态。
    """

    started_at = time.perf_counter()
    logger.info(
        "data query ingestion started doc_id=%s title=%s file_type=%s status=%s",
        document.doc_id,
        document.doc_title,
        document.file_type,
        document.status,
    )

    # 1. table_meta 是上传阶段的表名占位；worker 只消费属于当前 doc_id 的占位。
    stage_started_at = time.perf_counter()
    table_meta = await document_repository.get_table_meta_by_document(document_id=document.doc_id)
    if table_meta is None or table_meta.document_id != document.doc_id:
        raise DataQueryIngestionFailed()
    logger.info(
        "data query table meta loaded doc_id=%s namespace=%s table_name=%s elapsed_ms=%.2f",
        document.doc_id,
        table_meta.namespace,
        table_meta.table_name,
        (time.perf_counter() - stage_started_at) * 1000,
    )

    # 2. DATA_QUERY 使用原始上传文件作为数据源，不依赖 converted_doc_url。
    object_key = original_object_key(doc_id=document.doc_id, safe_filename=document.doc_title)
    stage_started_at = time.perf_counter()
    content = await storage.download_bytes(object_key=object_key)
    logger.info(
        "data query source downloaded doc_id=%s object_key=%s size_bytes=%s elapsed_ms=%.2f",
        document.doc_id,
        object_key,
        len(content),
        (time.perf_counter() - stage_started_at) * 1000,
    )

    # 3. 解析和计划构建都不落库，便于任何失败在进入事务前直接向外传播。
    stage_started_at = time.perf_counter()
    dataset = parse_data_query_spreadsheet(
        file_type=document.file_type,
        content=content,
        filename=document.doc_title,
    )
    logger.info(
        "data query spreadsheet parsed doc_id=%s sheet=%s columns=%s rows=%s elapsed_ms=%.2f",
        document.doc_id,
        dataset.original_sheet_name,
        len(dataset.headers),
        len(dataset.rows),
        (time.perf_counter() - stage_started_at) * 1000,
    )
    stage_started_at = time.perf_counter()
    plan = build_data_query_table_plan(
        namespace=table_meta.namespace,
        table_name=table_meta.table_name,
        dataset=dataset,
    )
    logger.info(
        "data query table plan built doc_id=%s physical_table=%s columns=%s rows=%s "
        "elapsed_ms=%.2f",
        document.doc_id,
        plan.physical_table_name,
        len(plan.column_names),
        len(plan.rows),
        (time.perf_counter() - stage_started_at) * 1000,
    )
    stage_started_at = time.perf_counter()
    await document_repository.import_data_query_table(
        document_id=document.doc_id,
        physical_table_name=plan.physical_table_name,
        create_sql=plan.create_sql,
        columns_info=plan.columns_info,
        column_names=plan.column_names,
        rows=plan.rows,
    )
    logger.info(
        "data query table imported doc_id=%s physical_table=%s columns=%s rows=%s "
        "db_elapsed_ms=%.2f total_elapsed_ms=%.2f",
        document.doc_id,
        plan.physical_table_name,
        len(plan.column_names),
        len(plan.rows),
        (time.perf_counter() - stage_started_at) * 1000,
        (time.perf_counter() - started_at) * 1000,
    )


def build_data_query_table_plan(
    *,
    namespace: str,
    table_name: str,
    dataset: ParsedDataQueryTable,
) -> DataQueryTablePlan:
    """根据解析结果生成动态表 identifier、DDL 和 metadata。

    物理列名统一使用 `col_001`、`col_002`，不直接使用用户表头。这样可以规避空表头、
    重复表头、中文表头、SQL 关键字和特殊字符等问题，同时在 `columns_info` 中保留
    原始表头供后续 Text2SQL 展示或 prompt 组装使用。
    """

    physical_table_name = build_physical_table_name(namespace=namespace, table_name=table_name)
    column_names = [f"col_{index:03d}" for index in range(1, len(dataset.headers) + 1)]
    # columns_info 是面向后续查询链路的“用户可读 schema”，不是数据库实际列名来源。
    columns_info = {
        "originalSheetName": dataset.original_sheet_name,
        "physicalTableName": physical_table_name,
        "columns": [
            {
                "ordinal": index,
                "header": header,
                "columnName": column_name,
                "type": "TEXT",
            }
            for index, (header, column_name) in enumerate(
                zip(dataset.headers, column_names, strict=True),
                start=1,
            )
        ],
    }
    return DataQueryTablePlan(
        physical_table_name=physical_table_name,
        column_names=column_names,
        create_sql=build_create_table_sql(
            physical_table_name=physical_table_name,
            column_names=column_names,
        ),
        columns_info=columns_info,
        rows=dataset.rows,
    )


def build_physical_table_name(*, namespace: str, table_name: str) -> str:
    """生成 PostgreSQL 安全的 DATA_QUERY 物理表名。"""

    return build_data_query_physical_table_name(namespace=namespace, table_name=table_name)


def build_create_table_sql(*, physical_table_name: str, column_names: list[str]) -> str:
    """构造动态建表 DDL。

    只有后端生成且已校验的表名、列名会被拼进 SQL；所有列当前统一使用 TEXT，
    第一版不做类型推断，避免上传数据变化导致 schema 不稳定。
    """

    quoted_table = _quote_identifier(physical_table_name)
    column_defs = ", ".join(f"{_quote_identifier(column_name)} TEXT" for column_name in column_names)
    return f"CREATE TABLE {quoted_table} ({column_defs})"


def build_insert_sql_and_params(
    *,
    physical_table_name: str,
    column_names: list[str],
    row: list[str],
) -> tuple[str, dict[str, str]]:
    """构造参数化 INSERT 语句及其行参数。

    SQL 字符串里只包含后端生成的 identifier；单元格内容全部放入 params，
    由 SQLAlchemy 绑定，避免把用户上传的表格内容拼接到 SQL 中。
    """

    quoted_table = _quote_identifier(physical_table_name)
    quoted_columns = ", ".join(_quote_identifier(column_name) for column_name in column_names)
    placeholders = ", ".join(f":{column_name}" for column_name in column_names)
    # 行宽不足时补空字符串，行宽过长时截断到表头宽度，保证 INSERT 列和值一一对应。
    values = list(row[: len(column_names)])
    values.extend([""] * (len(column_names) - len(values)))
    return (
        f"INSERT INTO {quoted_table} ({quoted_columns}) VALUES ({placeholders})",
        dict(zip(column_names, values, strict=True)),
    )


def build_insert_params(*, column_names: list[str], row: list[str]) -> dict[str, str]:
    """构造单行 INSERT 参数，供批量 execute 复用同一条 SQL。"""

    values = list(row[: len(column_names)])
    values.extend([""] * (len(column_names) - len(values)))
    return dict(zip(column_names, values, strict=True))


def _parse_excel(*, content: bytes, filename: str) -> ParsedDataQueryTable:
    """解析 Excel 并强制收敛为唯一数据 sheet。"""

    if str(filename).lower().endswith(".xls"):
        return _parse_xls(content)
    return _parse_xlsx(content)


def _parse_xlsx(content: bytes) -> ParsedDataQueryTable:
    """解析 xlsx 并强制收敛为唯一数据 sheet。"""

    try:
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise DataQuerySpreadsheetInvalid("invalid xlsx workbook") from exc
    data_tables: list[ParsedDataQueryTable] = []
    try:
        for worksheet in workbook.worksheets:
            # 1. 完全空 sheet 可忽略；非空但没有数据行的 header-only sheet 必须拒绝。
            rows = _worksheet_rows(worksheet)
            if not rows:
                continue
            header = [_cell_to_text(value) for value in rows[0]]
            data_rows = [_normalize_data_row(row, len(header)) for row in rows[1:]]
            data_rows = [row for row in data_rows if any(cell != "" for cell in row)]
            if not data_rows:
                raise DataQuerySpreadsheetInvalid("header-only sheet")
            # 2. 暂存有效数据 sheet，循环结束后统一判断是否超过一个。
            data_tables.append(
                ParsedDataQueryTable(
                    original_sheet_name=worksheet.title,
                    headers=header,
                    rows=data_rows,
                )
            )
    finally:
        workbook.close()

    if len(data_tables) > 1:
        raise DataQuerySpreadsheetInvalid("multiple data sheets")
    if not data_tables:
        raise DataQuerySpreadsheetInvalid("empty workbook")
    return data_tables[0]


def _parse_xls(content: bytes) -> ParsedDataQueryTable:
    """解析旧版 xls 并强制收敛为唯一数据 sheet。"""

    import xlrd

    try:
        workbook = xlrd.open_workbook(file_contents=content)
    except Exception as exc:
        raise DataQuerySpreadsheetInvalid("invalid xls workbook") from exc

    data_tables: list[ParsedDataQueryTable] = []
    for worksheet in workbook.sheets():
        rows = _xls_worksheet_rows(worksheet)
        if not rows:
            continue
        header = [_cell_to_text(value) for value in rows[0]]
        data_rows = [_normalize_data_row(row, len(header)) for row in rows[1:]]
        data_rows = [row for row in data_rows if any(cell != "" for cell in row)]
        if not data_rows:
            raise DataQuerySpreadsheetInvalid("header-only sheet")
        data_tables.append(
            ParsedDataQueryTable(
                original_sheet_name=worksheet.name,
                headers=header,
                rows=data_rows,
            )
        )

    if len(data_tables) > 1:
        raise DataQuerySpreadsheetInvalid("multiple data sheets")
    if not data_tables:
        raise DataQuerySpreadsheetInvalid("empty workbook")
    return data_tables[0]


def _parse_csv(content: bytes) -> ParsedDataQueryTable:
    """解析 CSV 为单表数据，并拒绝空文件或只有表头的文件。"""

    text = content.decode("utf-8-sig")
    rows = [list(row) for row in csv.reader(StringIO(text))]
    rows = [_trim_trailing_empty(row) for row in rows]
    rows = [row for row in rows if any(_cell_to_text(cell) != "" for cell in row)]
    if len(rows) < 2:
        raise DataQuerySpreadsheetInvalid("CSV requires data rows")
    headers = [_cell_to_text(value) for value in rows[0]]
    data_rows = [_normalize_data_row(row, len(headers)) for row in rows[1:]]
    data_rows = [row for row in data_rows if any(cell != "" for cell in row)]
    if not data_rows:
        raise DataQuerySpreadsheetInvalid("CSV requires data rows")
    return ParsedDataQueryTable(
        original_sheet_name="Data",
        headers=headers,
        rows=data_rows,
    )


def _worksheet_rows(worksheet) -> list[list[Any]]:
    """读取 worksheet 中所有非空行，并去掉每行末尾的空单元格。"""

    rows: list[list[Any]] = []
    for row in worksheet.iter_rows(values_only=True):
        trimmed = _trim_trailing_empty(list(row))
        if any(_cell_to_text(value) != "" for value in trimmed):
            rows.append(trimmed)
    return rows


def _xls_worksheet_rows(worksheet) -> list[list[Any]]:
    """读取 xls worksheet 中所有非空行，并去掉每行末尾的空单元格。"""

    rows: list[list[Any]] = []
    for row_index in range(worksheet.nrows):
        trimmed = _trim_trailing_empty(list(worksheet.row_values(row_index)))
        if any(_cell_to_text(value) != "" for value in trimmed):
            rows.append(trimmed)
    return rows


def _normalize_data_row(row: list[Any], width: int) -> list[str]:
    """把数据行规整到表头宽度并统一转为字符串。"""

    values = [_cell_to_text(value) for value in row[:width]]
    values.extend([""] * (width - len(values)))
    return values


def _trim_trailing_empty(row: list[Any]) -> list[Any]:
    """去掉行末尾连续空值，保留中间空单元格。"""

    trimmed = list(row)
    while trimmed and _cell_to_text(trimmed[-1]) == "":
        trimmed.pop()
    return trimmed


def _cell_to_text(value: Any) -> str:
    """把 Excel/CSV 单元格值转成入库使用的字符串。"""

    if value is None:
        return ""
    return str(value)


def _quote_identifier(identifier: str) -> str:
    """对后端生成的 identifier 做最终校验和 quote。"""

    return quote_generated_identifier(identifier)
