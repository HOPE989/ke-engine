"""文档切分、转换结果加载与 segment draft 构造能力。

本模块的边界是“已完成转换的文档如何进入分片阶段”：
- Markdown/PDF/Word 链路读取 converted Markdown，并按 Markdown header 做父子分片。
- Excel/CSV 链路读取 origin bytes，先转成紧凑 HTML table section，再按同一套父子分片语义输出。

workflow 不直接关心具体文件如何读取和解析，只负责按 file_type 选择 splitter。
"""

from collections.abc import Awaitable, Callable
import csv
from dataclasses import dataclass
import html
from io import BytesIO, StringIO
import logging
from typing import Any
from urllib.parse import unquote, urlparse

from langchain_text_splitters import MarkdownHeaderTextSplitter, RecursiveCharacterTextSplitter
from langchain_text_splitters.base import TextSplitter
from starlette.concurrency import run_in_threadpool

from app.modules.document.errors import (
    ChunkLockUnavailable,
    ChunkSplittingFailed,
    ConvertedMarkdownInvalid,
    ConvertedMarkdownUnavailable,
    DocumentStateConflict,
)
from app.modules.document.file_types import DocumentFileType
from app.modules.document.markdown import parse_markdown_image_references

HEADERS_TO_SPLIT_ON = [
    ("#", "Header 1"),
    ("##", "Header 2"),
    ("###", "Header 3"),
    ("####", "Header 4"),
    ("#####", "Header 5"),
    ("######", "Header 6"),
]

RECURSIVE_SEPARATORS = [
    "\n\n",
    "\n",
    " ",
    ".",
    ",",
    "\u200b",
    "\uff0c",
    "\u3001",
    "\uff0e",
    "\u3002",
    "",
]

logger = logging.getLogger(__name__)


def _file_type_value(file_type: DocumentFileType | str) -> str:
    if isinstance(file_type, DocumentFileType):
        return file_type.value
    return str(file_type)


@dataclass(frozen=True, slots=True)
class MarkdownSplitChunk:
    """文档 splitter 产出的待持久化分段。

    `skip_embedding=True` 表示这是一个父块，只入库用于召回后回填上下文，
    不发送向量化；子块通过 `parent_chunk_id` 指向父块并参与 embedding。
    """

    chunk_id: str
    text: str
    langchain_metadata: dict[str, Any]
    skip_embedding: bool
    parent_chunk_id: str | None


@dataclass(frozen=True, slots=True)
class SegmentDraft:
    """准备批量写入 knowledge_segment 的分段数据。"""

    id: int
    chunk_id: str
    text: str
    document_id: int
    chunk_order: int
    embedding_id: str | None
    status: str
    metadata: dict[str, Any]
    skip_embedding: bool


@dataclass(frozen=True, slots=True)
class HTMLTableSection:
    """Excel/CSV 预切分得到的紧凑 HTML table section。

    这里的 section 是“初次分片”的结果：一个 section 对应同一个 sheet 中
    第一行 header 加连续 N 行数据。后续再根据字符长度决定是否继续做父子分片。
    """

    text: str
    metadata: dict[str, Any]


class _LocalIdGenerator:
    def __init__(self) -> None:
        self._next_value = 0

    def next_id(self) -> int:
        self._next_value += 1
        return self._next_value


class MarkdownHeaderParentTextSplitter(TextSplitter):
    """按 Markdown 标题切父块，超长父块再按递归字符切子块。

    这个类同时保留 LangChain `TextSplitter.split_text()` 兼容入口，业务链路则统一
    调用 async `split_chunks()`：先从对象存储下载 converted Markdown，再进入纯文本切分。
    """

    def __init__(self, *, chunk_size: int, overlap: int) -> None:
        super().__init__(
            chunk_size=chunk_size,
            chunk_overlap=overlap,
            length_function=len,
        )

    def split_text(self, text: str) -> list[str]:
        """返回 LangChain TextSplitter 兼容的文本切分结果。"""

        return [
            chunk.text
            for chunk in self.split_markdown_chunks(
                text,
                id_generator=_LocalIdGenerator(),
            )
        ]

    async def split_chunks(
        self,
        *,
        document: Any,
        storage: Any,
        id_generator: Any,
    ) -> list[MarkdownSplitChunk]:
        """下载转换后的 Markdown，并返回业务持久化需要的 chunk 结构。

        下载与 UTF-8 decode 留在 splitter 内部，是为了让 workflow 对不同 file_type
        保持一致：Markdown splitter 读取文本，Excel splitter 读取原始 bytes。
        """

        markdown = await load_converted_markdown(document=document, storage=storage)
        return await run_in_threadpool(
            self.split_markdown_chunks,
            markdown,
            id_generator=id_generator,
        )

    def split_markdown_chunks(self, text: str, *, id_generator: Any) -> list[MarkdownSplitChunk]:
        """对已经加载好的 Markdown 文本执行父子分片。

        关键步骤：
        1. 先用 Markdown header 识别语义 section，header 本身进入 metadata。
        2. section 不超过 chunk_size 时，直接作为可 embedding chunk。
        3. section 超长时，完整 section 作为 skip parent，再递归切出 children 用于检索。
        """

        header_splitter = MarkdownHeaderTextSplitter(
            headers_to_split_on=HEADERS_TO_SPLIT_ON,
            strip_headers=True,
            return_each_line=False,
        )
        chunks: list[MarkdownSplitChunk] = []
        recursive_splitter: RecursiveCharacterTextSplitter | None = None

        for section in header_splitter.split_text(text):
            section_text = section.page_content
            if not section_text.strip():
                continue

            metadata = dict(section.metadata)
            # 短 section 语义足够完整，直接参与 embedding。
            if len(section_text) <= self._chunk_size:
                chunks.append(
                    MarkdownSplitChunk(
                        chunk_id=str(id_generator.next_id()),
                        text=section_text,
                        langchain_metadata=metadata,
                        skip_embedding=False,
                        parent_chunk_id=None,
                    )
                )
                continue

            if recursive_splitter is None:
                recursive_splitter = RecursiveCharacterTextSplitter(
                    chunk_size=self._chunk_size,
                    chunk_overlap=self._chunk_overlap,
                    length_function=len,
                    is_separator_regex=False,
                    separators=RECURSIVE_SEPARATORS,
                )
            # 长 section 保留完整父块，children 只负责检索命中，召回时再回填父块。
            parent_chunk_id = str(id_generator.next_id())
            chunks.append(
                MarkdownSplitChunk(
                    chunk_id=parent_chunk_id,
                    text=section_text,
                    langchain_metadata=metadata,
                    skip_embedding=True,
                    parent_chunk_id=None,
                )
            )
            for child_text in recursive_splitter.split_text(section_text):
                if not child_text.strip():
                    continue
                chunks.append(
                    MarkdownSplitChunk(
                        chunk_id=str(id_generator.next_id()),
                        text=child_text,
                        langchain_metadata=metadata,
                        skip_embedding=False,
                        parent_chunk_id=parent_chunk_id,
                    )
                )

        return chunks


class Excel2HTMLParentTextSplitter:
    """把 Excel/CSV 转成紧凑 HTML table section，再按父子块规则切分。

    设计目标是服务 DOCUMENT_SEARCH 的复杂表格检索，而不是结构化查询：
    - converter 阶段只回填 origin URL，不生成中间 Markdown。
    - splitter 阶段下载原始 bytes，并按 sheet 转为 RAGFlow-like HTML table section。
    - 每个 section 使用 `filename - sheetName` 作为 caption，补足文件名里携带的业务语义。
    - 第一行视为 header，每 12 行数据组成一个 section，section 内重复 header。
    """

    chunk_rows = 12

    def __init__(self, *, chunk_size: int, overlap: int) -> None:
        self._chunk_size = chunk_size
        self._chunk_overlap = overlap

    async def split_chunks(
        self,
        *,
        document: Any,
        storage: Any,
        id_generator: Any,
    ) -> list[MarkdownSplitChunk]:
        """下载原始表格文件，并返回业务持久化需要的 chunk 结构。

        Excel/CSV 的 `converted_doc_url` 指向 origin URL，因此这里必须读取 bytes，
        不能沿用 Markdown 的 UTF-8 decode 逻辑。
        """

        content = await load_converted_bytes(document=document, storage=storage)
        return await run_in_threadpool(
            self.split_table_bytes,
            document=document,
            content=content,
            id_generator=id_generator,
        )

    def split_table_bytes(
        self,
        *,
        document: Any,
        content: bytes,
        id_generator: Any,
    ) -> list[MarkdownSplitChunk]:
        """把表格 bytes 转为 HTML section，并继续执行 parent/child 分片。

        关键步骤：
        1. 解析 workbook/CSV，按 sheet 产出紧凑 HTML table section。
        2. section 小于等于 chunk_size 时，作为普通 chunk 参与 embedding。
        3. section 超过 chunk_size 时，完整 HTML section 作为 skip parent。
        4. children 直接由 RecursiveCharacterTextSplitter 切出，不额外补 caption/header。
        """

        # 先完成 Excel2HTML 初次分片。这里的 section 仍保持完整 HTML table 语义。
        sections = self._table_sections(document=document, content=content)
        chunks: list[MarkdownSplitChunk] = []
        recursive_splitter: RecursiveCharacterTextSplitter | None = None

        for section in sections:
            # 短 section 直接入向量库；metadata 记录 sheet、行号和 section 序号。
            if len(section.text) <= self._chunk_size:
                chunks.append(
                    MarkdownSplitChunk(
                        chunk_id=str(id_generator.next_id()),
                        text=section.text,
                        langchain_metadata=dict(section.metadata),
                        skip_embedding=False,
                        parent_chunk_id=None,
                    )
                )
                continue

            if recursive_splitter is None:
                recursive_splitter = RecursiveCharacterTextSplitter(
                    chunk_size=self._chunk_size,
                    chunk_overlap=self._chunk_overlap,
                    length_function=len,
                    is_separator_regex=False,
                    separators=RECURSIVE_SEPARATORS,
                )
            # 长 section 的完整 HTML table 是父块，保证召回上下文不丢表格结构。
            parent_chunk_id = str(id_generator.next_id())
            chunks.append(
                MarkdownSplitChunk(
                    chunk_id=parent_chunk_id,
                    text=section.text,
                    langchain_metadata=dict(section.metadata),
                    skip_embedding=True,
                    parent_chunk_id=None,
                )
            )
            for child_text in recursive_splitter.split_text(section.text):
                if not child_text.strip():
                    continue
                # 子块只用于检索命中，沿用父块 metadata，不重复拼接 caption/header。
                chunks.append(
                    MarkdownSplitChunk(
                        chunk_id=str(id_generator.next_id()),
                        text=child_text,
                        langchain_metadata=dict(section.metadata),
                        skip_embedding=False,
                        parent_chunk_id=parent_chunk_id,
                    )
                )

        return chunks

    def _table_sections(self, *, document: Any, content: bytes) -> list[HTMLTableSection]:
        """根据 document.file_type 解析原始内容，并收集所有 sheet 的 HTML section。"""

        file_type = _file_type_value(getattr(document, "file_type", ""))
        if file_type == DocumentFileType.CSV.value:
            # CSV 没有 sheet 概念，统一使用 Data 作为稳定 sheetName。
            sheets = [("Data", _csv_rows(content))]
        else:
            sheets = _excel_sheets(content=content, filename=getattr(document, "doc_title", ""))

        file_name = str(getattr(document, "doc_title", "") or "document")
        sections: list[HTMLTableSection] = []
        for sheet_name, rows in sheets:
            sections.extend(
                _rows_to_html_sections(
                    file_name=file_name,
                    sheet_name=sheet_name,
                    rows=rows,
                    chunk_rows=self.chunk_rows,
                    start_index=len(sections),
                )
            )
        return sections


def _csv_rows(content: bytes) -> list[tuple[int, list[str]]]:
    """读取 CSV bytes，返回非空行及其原始行号。"""

    text = _decode_csv_text(content)
    rows: list[tuple[int, list[str]]] = []
    for row_number, row in enumerate(csv.reader(StringIO(text)), start=1):
        cells = [_cell_text(cell) for cell in row]
        if _is_empty_row(cells):
            continue
        rows.append((row_number, cells))
    return rows


def _decode_csv_text(content: bytes) -> str:
    """按常见中文 CSV 编码顺序解码。

    `utf-8-sig` 处理带 BOM 的导出文件，`gb18030` 覆盖常见 Windows/Excel 中文 CSV，
    最后的 `latin-1` 用于保底避免直接解码失败。
    """

    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _excel_sheets(*, content: bytes, filename: str) -> list[tuple[str, list[tuple[int, list[str]]]]]:
    """按扩展名选择 xlsx/xls 解析器，返回 sheet 名和非空行。"""

    if str(filename).lower().endswith(".xls"):
        return _xls_sheets(content)
    return _xlsx_sheets(content)


def _xlsx_sheets(content: bytes) -> list[tuple[str, list[tuple[int, list[str]]]]]:
    """用 openpyxl 读取 xlsx，每个 sheet 独立容错。"""

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    try:
        sheets: list[tuple[str, list[tuple[int, list[str]]]]] = []
        for worksheet in workbook.worksheets:
            try:
                rows: list[tuple[int, list[str]]] = []
                for row_number, row in enumerate(
                    worksheet.iter_rows(values_only=True),
                    start=1,
                ):
                    cells = [_cell_text(cell) for cell in row]
                    if _is_empty_row(cells):
                        continue
                    rows.append((row_number, cells))
                sheets.append((worksheet.title, rows))
            except Exception:
                logger.warning(
                    "excel sheet parse failed",
                    extra={"sheet_name": getattr(worksheet, "title", None)},
                )
        return sheets
    finally:
        workbook.close()


def _xls_sheets(content: bytes) -> list[tuple[str, list[tuple[int, list[str]]]]]:
    """用 xlrd 读取旧版 xls，每个 sheet 独立容错。"""

    import xlrd

    workbook = xlrd.open_workbook(file_contents=content)
    sheets: list[tuple[str, list[tuple[int, list[str]]]]] = []
    for worksheet in workbook.sheets():
        try:
            rows: list[tuple[int, list[str]]] = []
            for row_index in range(worksheet.nrows):
                cells = [_cell_text(cell) for cell in worksheet.row_values(row_index)]
                if _is_empty_row(cells):
                    continue
                rows.append((row_index + 1, cells))
            sheets.append((worksheet.name, rows))
        except Exception:
            logger.warning(
                "excel sheet parse failed",
                extra={"sheet_name": getattr(worksheet, "name", None)},
            )
    return sheets


def _rows_to_html_sections(
    *,
    file_name: str,
    sheet_name: str,
    rows: list[tuple[int, list[str]]],
    chunk_rows: int,
    start_index: int,
) -> list[HTMLTableSection]:
    """把一个 sheet 的非空行转换为多个紧凑 HTML table section。

    第一条非空行固定作为 header；如果 sheet 只有 header 或为空，则不产出 section。
    每个 section 包含 caption、重复 header 和最多 `chunk_rows` 行数据。
    """

    if len(rows) <= 1:
        return []

    header_row_number, header_cells = rows[0]
    data_rows = rows[1:]
    sections: list[HTMLTableSection] = []
    caption = f"{file_name} - {sheet_name}"
    for offset in range(0, len(data_rows), chunk_rows):
        row_group = data_rows[offset : offset + chunk_rows]
        if not row_group:
            continue
        # caption 使用文件名和 sheet 名组合，保留“5月销售总额.xlsx”这类文件名语义。
        html_table = _html_table(
            caption=caption,
            header_cells=header_cells,
            data_rows=[cells for _, cells in row_group],
        )
        sections.append(
            HTMLTableSection(
                text=html_table,
                metadata={
                    "sourceFormat": "html_table",
                    "sheetName": sheet_name,
                    "headerRow": header_row_number,
                    "dataStartRow": row_group[0][0],
                    "dataEndRow": row_group[-1][0],
                    "chunkRows": chunk_rows,
                    "htmlTableIndex": start_index + len(sections),
                },
            )
        )
    return sections


def _html_table(*, caption: str, header_cells: list[str], data_rows: list[list[str]]) -> str:
    """生成无多余空白的 HTML table，降低无意义 token 和向量噪声。"""

    return (
        f"<table><caption>{_escape_html(caption)}</caption>"
        f"{_html_row('th', header_cells)}"
        f"{''.join(_html_row('td', row) for row in data_rows)}</table>\n"
    )


def _html_row(tag: str, values: list[str]) -> str:
    cells = "".join(f"<{tag}>{_escape_html(value)}</{tag}>" for value in values)
    return f"<tr>{cells}</tr>"


def _escape_html(value: str) -> str:
    return html.escape(value, quote=True)


def _cell_text(value: Any) -> str:
    """把单元格值归一化为稳定文本，避免 100.0 这类数值噪声。"""

    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _is_empty_row(cells: list[str]) -> bool:
    return not any(cell.strip() for cell in cells)


class DocumentSplitterFactory:
    """按文件类型唯一选择文档切分器。"""

    def __init__(self) -> None:
        self._builders: dict[str, Callable[..., Any]] = {}

    def register(
        self,
        *,
        file_type: DocumentFileType | str,
        splitter_builder: Callable[..., Any],
    ) -> None:
        normalized_file_type = _file_type_value(file_type)
        if normalized_file_type in self._builders:
            raise ValueError(f"splitter for file_type {normalized_file_type} already registered")
        self._builders[normalized_file_type] = splitter_builder

    def splitter_for(
        self,
        file_type: DocumentFileType | str,
        *,
        chunk_size: int,
        overlap: int,
    ) -> Any:
        builder = self._builders.get(_file_type_value(file_type))
        if builder is None:
            raise ChunkSplittingFailed()
        return builder(chunk_size=chunk_size, overlap=overlap)


def create_default_document_splitter_factory() -> DocumentSplitterFactory:
    """创建进程启动期使用的默认 splitter 注册表。"""

    factory = DocumentSplitterFactory()
    for file_type in (
        DocumentFileType.PLAIN_TEXT,
        DocumentFileType.PDF,
        DocumentFileType.WORD,
    ):
        factory.register(
            file_type=file_type,
            splitter_builder=MarkdownHeaderParentTextSplitter,
        )
    for file_type in (
        DocumentFileType.EXCEL,
        DocumentFileType.CSV,
    ):
        factory.register(
            file_type=file_type,
            splitter_builder=Excel2HTMLParentTextSplitter,
        )
    return factory


async def run_with_document_chunk_lock(
    *,
    lock: Any,
    operation: Callable[[], Awaitable[Any]],
) -> Any:
    """持有单文档切分锁执行一个异步操作。"""

    try:
        acquired = lock.acquire(blocking=False)
    except Exception as exc:
        raise ChunkLockUnavailable() from exc

    if not acquired:
        raise DocumentStateConflict()

    try:
        return await operation()
    finally:
        try:
            lock.release()
        except Exception:
            logger.exception("failed to release document chunk lock")


def _resolve_converted_object_key(*, converted_doc_url: str, storage: Any, doc_id: int) -> str:
    base = urlparse(storage.public_base_url.rstrip("/"))
    url = urlparse(converted_doc_url)
    if url.scheme != base.scheme or url.netloc != base.netloc:
        raise DocumentStateConflict()

    base_path = base.path.rstrip("/")
    url_path = unquote(url.path)
    if base_path:
        if url_path != base_path and not url_path.startswith(f"{base_path}/"):
            raise DocumentStateConflict()
        relative_path = url_path[len(base_path) :].lstrip("/")
    else:
        relative_path = url_path.lstrip("/")

    bucket, separator, object_key = relative_path.partition("/")
    if bucket != storage.bucket or not separator or not object_key.strip("/"):
        raise DocumentStateConflict()
    if not object_key.startswith(f"documents/{doc_id}/"):
        raise DocumentStateConflict()
    return object_key


async def load_converted_markdown(*, document: Any, storage: Any) -> str:
    """解析 converted_doc_url，下载并返回 UTF-8 Markdown 文本。"""

    content = await load_converted_bytes(document=document, storage=storage)
    try:
        return content.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise ConvertedMarkdownInvalid() from exc


async def load_converted_bytes(*, document: Any, storage: Any) -> bytes:
    """解析 converted_doc_url，下载并返回原始 bytes。"""

    converted_doc_url = getattr(document, "converted_doc_url", None)
    if not converted_doc_url:
        raise DocumentStateConflict()

    object_key = _resolve_converted_object_key(
        converted_doc_url=converted_doc_url,
        storage=storage,
        doc_id=document.doc_id,
    )
    try:
        content = await storage.download_bytes(object_key=object_key)
    except Exception as exc:
        raise ConvertedMarkdownUnavailable() from exc
    return content


def split_markdown_into_chunks(
    markdown: str,
    *,
    chunk_size: int,
    overlap: int,
    id_generator: Any,
) -> list[MarkdownSplitChunk]:
    """按 Markdown header 和递归字符长度切分 Markdown。"""

    return MarkdownHeaderParentTextSplitter(
        chunk_size=chunk_size,
        overlap=overlap,
    ).split_markdown_chunks(markdown, id_generator=id_generator)


def build_segment_drafts(
    *,
    document: Any,
    split_chunks: list[MarkdownSplitChunk],
    id_generator: Any,
) -> list[SegmentDraft]:
    """把 splitter 输出转换为可持久化的 segment drafts。"""

    drafts: list[SegmentDraft] = []
    for chunk_order, split_chunk in enumerate(split_chunks):
        row_id = id_generator.next_id()
        chunk_id = split_chunk.chunk_id
        metadata = {
            "skipEmbedding": split_chunk.skip_embedding,
            "chunkId": chunk_id,
            "docId": str(document.doc_id),
            "fileName": document.doc_title,
            "url": document.converted_doc_url,
            "accessibleBy": document.accessible_by,
            "parentChunkId": split_chunk.parent_chunk_id,
            "langchain": dict(split_chunk.langchain_metadata),
            "images": _extract_markdown_images(
                text=split_chunk.text,
                document=document,
            ),
        }
        drafts.append(
            SegmentDraft(
                id=row_id,
                chunk_id=chunk_id,
                text=split_chunk.text,
                document_id=document.doc_id,
                chunk_order=chunk_order,
                embedding_id=None,
                status="STORED",
                metadata=metadata,
                skip_embedding=split_chunk.skip_embedding,
            )
        )

    return drafts


def _extract_markdown_images(*, text: str, document: Any) -> list[dict[str, str]]:
    images: list[dict[str, str]] = []
    for reference in parse_markdown_image_references(text):
        image = {
            "url": reference.target,
            "alt": reference.alt,
            "source": "markdown-image",
        }
        object_key = _derive_document_image_object_key(
            image_url=reference.target,
            document=document,
        )
        if object_key is not None:
            image["objectKey"] = object_key
        images.append(image)
    return images


def _derive_document_image_object_key(*, image_url: str, document: Any) -> str | None:
    converted_doc_url = getattr(document, "converted_doc_url", None)
    if not converted_doc_url:
        return None

    converted = urlparse(converted_doc_url)
    image = urlparse(image_url)
    if image.scheme != converted.scheme or image.netloc != converted.netloc:
        return None

    doc_prefix = ["documents", str(document.doc_id)]
    converted_parts = [part for part in unquote(converted.path).split("/") if part]
    image_parts = [part for part in unquote(image.path).split("/") if part]
    marker_index = _find_subsequence(converted_parts, doc_prefix)
    if marker_index is None or marker_index == 0:
        return None

    base_and_bucket_parts = converted_parts[:marker_index]
    if image_parts[: len(base_and_bucket_parts)] != base_and_bucket_parts:
        return None

    object_key_parts = image_parts[len(base_and_bucket_parts) :]
    if object_key_parts[:2] != doc_prefix:
        return None

    return "/".join(object_key_parts)


def _find_subsequence(parts: list[str], subsequence: list[str]) -> int | None:
    for index in range(0, len(parts) - len(subsequence) + 1):
        if parts[index : index + len(subsequence)] == subsequence:
            return index
    return None
